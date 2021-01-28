import openpyxl

path = "C:/Users/Dell/Desktop/music.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
rows = sheet.max_row
cols = sheet.max_column

for row in range(1, rows + 1):
    for col in range(1, cols + 1):
        print(sheet.cell(row = row, column= col).value, end="     ")
    print()
